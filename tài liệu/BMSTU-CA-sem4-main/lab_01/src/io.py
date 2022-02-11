import openpyxl as xls

def parse_table():
    '''
        Загрузка таблицы в программу.
    '''
    pos = 3
    points = xls.load_workbook("points.xlsx").active
    table = []
    while points.cell(row = pos, column = 1).value is not None:
        table.append([float(points.cell(row = pos, column = 1).value), \
                      float(points.cell(row = pos, column = 2).value), \
                      float(points.cell(row = pos, column = 3).value)])
        pos += 1
    table.sort()
    print(table)
    count = len(table) - 1
    arr_x = []
    arr_y = []
    for i in range(len(table)):
        arr_x.append(table[i][0])
        arr_y.append(table[i][1])
    return table, arr_x, arr_y, count


def input_x():
    '''
        Ввод аргумента. (в случае ошибки дается еще попытка)
    '''
    print("Enter X: ")
    flag = 0
    x = 0
    while flag == 0:
        x = input(float)
        try:
            val = float(x)
            flag = 1
        except ValueError:
            print("Some error! Try again")
    return float(x)